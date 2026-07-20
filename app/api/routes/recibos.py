from fastapi import APIRouter
from openpyxl import load_workbook
from copy import copy

router = APIRouter(prefix="/recibos", tags=["recibos"])

wb = load_workbook("templates/RECIBO INMOBILIARIO.xlsx")

@router.get("/")
def get_recibos() -> dict:
    ws=wb.sheetnames
    return{
        "cantidad_hojas": len(ws),
        "hojas": ws
    }

@router.get("/actualizar")
def actualizar_fecha() -> dict:
        actualizar =get_recibos_ajustar() #trae recibos a actualizar
        print(actualizar)
        #recorre todas las hojas del excel
        for recibo in wb.sheetnames:
            ws = wb[recibo]
            #cambia la fecha del recibo y el mes
            ws["I13"].value = 8
            ws["C22"].value = "AGOSTO"

            if recibo in actualizar: #si tiene q actualizar 
                estilo = copy(ws["E22"]._style) #copia formato de la celda
                ws["E22"].value = 100000  #asigna el valor
                ws["E22"]._style = estilo #pega formato de celda
                print("ENTRO")
        wb.save("templates/RECIBO INMOBILIARIO.xlsx")
        print("READY")
        return {
            "cantidad_hojas": "LISTO",
        } 

def get_recibos_ajustar() -> list:
    recibos = ["00000(2)","00000(4)"]
   
    return recibos